#!/usr/bin/env python3
"""
LIQO Excel → D1 importer (parser stage).

Reads the four operational exports and emits SQL that matches schema.sql exactly,
plus a raw-inventory JSON for the engine-backed inventory stage (inventory_sql.ts).

    Sales Person List      → employees                (+ stores referenced)
    Stock Status (MC-wise) → inventory_raw.json       → 03_inventory.sql (engine)
    100 Cx Retails    ┐
    Last month Zirk   ┘    → customers, sales, sale_items, customer_events,
                             customer_brand_prefs       (grouped by bill)

PRIVACY: customers/sales/sale_items/customer_events carry PII (names, phones,
addresses) and dealer pricing. The output dir (data/import/) is gitignored and
must NEVER be committed — load it straight into your private D1 with wrangler.

Usage:
    python3 tools/etl/parse.py --in <uploads_dir> --out data/import
Files are matched by name substring, so your own re-exports work too.
"""
import argparse, json, os, re, glob, datetime as dt

# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def slug(s: str) -> str:
    return re.sub(r"-+$", "", re.sub(r"^-+", "", re.sub(r"[^a-z0-9]+", "-", (s or "").strip().lower())))

def q(v) -> str:
    """SQL string literal or NULL."""
    if v is None:
        return "NULL"
    s = str(v).strip()
    if s == "" or s.lower() == "none":
        return "NULL"
    return "'" + s.replace("'", "''") + "'"

def qi(v) -> str:
    """SQL integer literal or NULL."""
    if v is None or v == "":
        return "NULL"
    try:
        return str(int(round(float(v))))
    except (TypeError, ValueError):
        return "NULL"

def qf(v) -> str:
    if v is None or v == "":
        return "NULL"
    try:
        return repr(float(v))
    except (TypeError, ValueError):
        return "NULL"

def norm_phone(v):
    """Return a clean 10-digit Indian mobile, or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    d = re.sub(r"\D", "", s)
    if len(d) > 10 and d.startswith("91"):
        d = d[2:]
    if len(d) > 10 and d.startswith("0"):
        d = d[-10:]
    if len(d) == 10 and d[0] in "6789":
        return d
    return None

def to_ts(v):
    """Cell → 'YYYY-MM-DDT00:00:00' or None."""
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%dT%H:%M:%S") if isinstance(v, dt.datetime) else v.strftime("%Y-%m-%dT00:00:00")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%b-%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%dT00:00:00")
        except ValueError:
            continue
    return None

def first_num(v):
    if v is None:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", str(v))
    return float(m.group(0)) if m else None

def clean_name(name: str) -> str:
    """Drop a trailing employee code embedded in the name (e.g. 'Anjali LIQO1832')."""
    return re.sub(r"\s*LIQO\d+\s*$", "", (name or "").strip()).strip()

# engine category from a free-text product category (only the 4 recommendable ones)
def inv_cat(pc: str):
    s = (pc or "").strip().lower()
    if s in ("led tv", "tv", "television", "smart tv", "oled tv", "qled tv"):
        return "tv"
    if "air condition" in s or s in ("ac", "split ac", "window ac", "cassette ac", "tower ac"):
        return "ac"
    if "refriger" in s or s in ("fridge", "deep freezer"):
        return "fridge" if "refriger" in s or s == "fridge" else None  # deep freezer ≠ fridge engine cat
    if "washing" in s or s in ("washing machine", "washer", "washer dryer"):
        return "wm"
    return None

# sale-line analytics category: align the 4 engine slugs, else a general slug
def line_cat(itemgroup: str, main: str):
    s = (itemgroup or "").strip().lower()
    if "air cond" in s:
        return "ac"
    if "led tv" in s or s == "tv" or "television" in s:
        return "tv"
    if "refriger" in s or s == "fridge":
        return "fridge"
    if "washing" in s:
        return "wm"
    return slug(main or itemgroup or "other") or "other"

def premium_tier(ltv: float) -> str:
    if ltv >= 200000:
        return "luxury"
    if ltv >= 100000:
        return "premium"
    if ltv >= 40000:
        return "mainstream"
    return "value"

STORE_NAMES = {
    "zirakpur": "Liqo Zirakpur", "panchkula": "Liqo Panchkula", "chandigarh": "Liqo Chandigarh",
    "kharar": "Liqo Kharar", "pinjore": "Liqo Pinjore", "solan": "Liqo Solan",
    "ramgarh": "Liqo Ramgarh", "garments-warehouse": "Garments Warehouse",
}
# Stock-status per-store stock columns: header index → (store name, channel)
INV_STORE_COLS = {
    21: ("Chandigarh", "retail"), 23: ("Garments Warehouse", "logistics"),
    25: ("Kharar", "retail"), 27: ("Panchkula", "retail"), 29: ("Pinjore", "retail"),
    31: ("Ramgarh", "B2B"), 33: ("Ramgarh", "retail"), 35: ("Ramgarh", "QC"),
    37: ("Solan", "retail"), 39: ("Zirakpur", "retail"),
}

# When consent flag is set, imported historical customers are recallable in-app.
# Basis: first-party transaction records the retailer already holds. Set 0 to
# import contactless (no recall) — pass --no-consent.
def main():
    import openpyxl
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="indir", required=True)
    ap.add_argument("--out", dest="outdir", default="data/import")
    ap.add_argument("--no-consent", action="store_true", help="store customers with consent=0 (not recallable)")
    args = ap.parse_args()
    consent = 0 if args.no_consent else 1

    def find(sub):
        hits = glob.glob(os.path.join(args.indir, f"*{sub}*"))
        if not hits:
            raise SystemExit(f"missing input matching *{sub}*")
        return sorted(hits)[0]

    os.makedirs(os.path.join(args.outdir, "raw"), exist_ok=True)
    stores_used = set()
    summary = {}

    # ---------------------------------------------------------------- employees
    f_emp = find("Sales_Person")
    wb = openpyxl.load_workbook(f_emp, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    emp_ids, emp_sql = set(), []
    for r in rows:
        eid = (str(r[1]).strip() if len(r) > 1 and r[1] else "")
        if not eid or eid in emp_ids:
            continue
        emp_ids.add(eid)
        base = slug(r[0]) if len(r) > 0 else ""
        if base:
            stores_used.add(base)
        name = clean_name(r[2] if len(r) > 2 else "") or eid
        phone = norm_phone(r[4] if len(r) > 4 else None)
        desig = (str(r[5]).strip() if len(r) > 5 and r[5] else "")
        staus = (str(r[3]).strip() if len(r) > 3 and r[3] else "")
        role = "manager" if desig.lower() == "manager" else "salesperson"
        title = " · ".join([x for x in (desig, staus) if x]) or None
        emp_sql.append(
            f"INSERT OR IGNORE INTO employees (employee_id,name,phone,role,store_id,title,status) "
            f"VALUES ({q(eid)},{q(name)},{q(phone)},{q(role)},{q(base or None)},{q(title)},'active');"
        )
    summary["employees"] = len(emp_sql)

    # ---------------------------------------------------------------- inventory raw
    f_inv = find("StockStatus")
    wb = openpyxl.load_workbook(f_inv, read_only=True, data_only=True)
    ws = wb["ALL Inv."]
    it = ws.iter_rows(min_row=8, values_only=True)
    next(it)  # header row 8
    inv_raw, cat_skip = [], {}
    for r in rows_iter(it):
        pc = r[1]
        cat = inv_cat(pc)
        if not cat:
            cat_skip[str(pc).strip()] = cat_skip.get(str(pc).strip(), 0) + 1
            continue
        sku = (str(r[0]).strip() if r[0] else "")
        if not sku:
            continue
        desc_blob = " ".join(str(r[i]) for i in (2, 3, 4, 5, 9) if len(r) > i and r[i]).lower()
        star = None
        m = re.search(r"([1-5])\s*star", desc_blob)
        if m:
            star = int(m.group(1))
        inverter = "inverter" in desc_blob
        type2 = (str(r[13]).strip() if len(r) > 13 and r[13] else "")
        smart = type2 if any(k in type2.lower() for k in ("tv", "vidaa", "google", "android", "web", "tizen", "fire", "roku", "coolita")) else None
        price = None
        for ci in (19, 18, 16):  # LIQO, ONLINE, MRP
            v = r[ci] if len(r) > ci else None
            if v and float(v) > 0:
                price = float(v)
                break
        mrp = r[16] if len(r) > 16 and r[16] else None
        dealer = r[17] if len(r) > 17 and r[17] else None
        sku_margin = None
        if price and dealer and float(price) > float(dealer) > 0:
            sku_margin = round(float(price) - float(dealer))
        cap = first_num(r[10] if len(r) > 10 else None)
        for ci, (store, channel) in INV_STORE_COLS.items():
            qty = r[ci] if len(r) > ci else None
            try:
                qn = int(round(float(qty)))
            except (TypeError, ValueError):
                qn = 0
            if qn <= 0:
                continue
            stores_used.add(slug(store))
            inv_raw.append({
                "sku": sku, "store": store, "channel": channel,
                "category": cat, "category_label": str(pc).strip() if pc else None,
                "brand": (str(r[8]).strip() if len(r) > 8 and r[8] else None),
                "model": (str(r[7]).strip() if len(r) > 7 and r[7] else None),
                "name": (str(r[2]).strip() if len(r) > 2 and r[2] else None),
                "subCategory": (str(r[9]).strip() if len(r) > 9 and r[9] else None),
                "capacityValue": cap, "capacityText": (str(r[10]).strip() if len(r) > 10 and r[10] else None),
                "starRating": star, "inverter": inverter, "smartOS": smart,
                "price": round(price) if price else None,
                "mrp": round(float(mrp)) if mrp else None,
                "qty": qn, "stockQty": qn, "skuMargin": sku_margin,
            })
    wb.close()
    with open(os.path.join(args.outdir, "raw", "inventory_raw.json"), "w") as fh:
        json.dump(inv_raw, fh)
    summary["inventory_raw_rows"] = len(inv_raw)
    summary["inventory_categories_skipped"] = cat_skip

    # ---------------------------------------------------------------- sales (2 files)
    bills = {}            # bill_no → dict
    cust = {}             # phone → dict
    for sub in ("Cx_Retails", "Last_month_Sales"):
        path = find(sub)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        sit = ws.iter_rows(min_row=2, values_only=True)
        for r in rows_iter(sit):
            bill_no = (str(r[5]).strip() if len(r) > 5 and r[5] else "")
            if not bill_no:
                continue
            phone = norm_phone(r[7] if len(r) > 7 else None)
            loc = slug(r[1]) if len(r) > 1 else ""
            if loc:
                stores_used.add(loc)
            ref = (str(r[11]).strip() if len(r) > 11 and r[11] else "")
            mref = re.search(r"LIQO\d+", ref)
            emp = mref.group(0) if mref and mref.group(0) in emp_ids else None
            ts = to_ts(r[4] if len(r) > 4 else None)
            qty = first_num(r[26]) or 1
            item_amt = first_num(r[27]) or 0
            total_amt = first_num(r[30]) or item_amt
            brand = (str(r[20]).strip() if len(r) > 20 and r[20] else None)
            cat = line_cat(r[12] if len(r) > 12 else "", r[39] if len(r) > 39 else "")
            line = {
                "sku": (str(r[19]).strip() if len(r) > 19 and r[19] else (str(r[13]).strip() if len(r) > 13 and r[13] else None)),
                "brand": brand, "category": cat, "qty": int(qty),
                "unit_price": round(item_amt / qty) if qty else round(item_amt),
                "line_total": round(total_amt),
            }
            b = bills.get(bill_no)
            if not b:
                b = bills[bill_no] = {"phone": phone, "store": loc, "emp": emp, "ts": ts, "lines": []}
            b["lines"].append(line)
            if phone:
                c = cust.get(phone)
                if not c:
                    c = cust[phone] = {
                        "name": (str(r[6]).strip() if len(r) > 6 and r[6] else None),
                        "store": loc, "first": ts, "last": ts, "ltv": 0, "brands": set(),
                    }
                c["ltv"] += round(total_amt)
                if brand:
                    c["brands"].add(brand)
                if ts and (not c["first"] or ts < c["first"]):
                    c["first"] = ts
                if ts and (not c["last"] or ts > c["last"]):
                    c["last"] = ts
        wb.close()

    # customers + brand prefs
    cust_sql, bp_sql = [], []
    bp_id = 0
    for phone, c in cust.items():
        cid = "c-" + phone
        cust_sql.append(
            f"INSERT OR IGNORE INTO customers (customer_id,phone,name,consent,premium_tier,home_store_id,first_seen_at,last_seen_at) "
            f"VALUES ({q(cid)},{q(phone)},{q(c['name'])},{consent},{q(premium_tier(c['ltv']))},{q(c['store'] or None)},{q(c['first'])},{q(c['last'])});"
        )
        for br in sorted(c["brands"]):
            bp_id += 1
            bp_sql.append(
                f"INSERT OR IGNORE INTO customer_brand_prefs (id,customer_id,brand,affinity) "
                f"VALUES ({bp_id},{q(cid)},{q(br)},'owns');"
            )
    summary["customers"] = len(cust_sql)
    summary["brand_prefs"] = len(bp_sql)

    # sales + items + events
    sales_sql, item_sql, ev_sql = [], [], []
    sid = it_id = ev_id = 0
    for bill_no, b in bills.items():
        sid += 1
        cid = "c-" + b["phone"] if b["phone"] else None
        total = sum(l["line_total"] for l in b["lines"])
        items_count = sum(l["qty"] for l in b["lines"])
        sales_sql.append(
            f"INSERT OR IGNORE INTO sales (sale_id,bill_no,customer_id,employee_id,store_id,total,items_count,source,ts) "
            f"VALUES ({sid},{q(bill_no)},{q(cid)},{q(b['emp'])},{q(b['store'] or None)},{total},{items_count},'walk_in',{q(b['ts'] or '1970-01-01T00:00:00')});"
        )
        for l in b["lines"]:
            it_id += 1
            item_sql.append(
                f"INSERT OR IGNORE INTO sale_items (id,sale_id,sku,brand,category,qty,unit_price,line_total,recommended) "
                f"VALUES ({it_id},{sid},{q(l['sku'])},{q(l['brand'])},{q(l['category'])},{l['qty']},{l['unit_price']},{l['line_total']},0);"
            )
        if cid:
            primary = max(b["lines"], key=lambda x: x["line_total"])
            ev_id += 1
            ev_sql.append(
                f"INSERT OR IGNORE INTO customer_events (event_id,customer_id,type,category,brand,store_id,amount,ts) "
                f"VALUES ({ev_id},{q(cid)},'purchase',{q(primary['category'])},{q(primary['brand'])},{q(b['store'] or None)},{total},{q(b['ts'] or '1970-01-01T00:00:00')});"
            )
    summary["sales"] = len(sales_sql)
    summary["sale_items"] = len(item_sql)
    summary["customer_events"] = len(ev_sql)

    # stores (every slug referenced anywhere)
    store_sql = []
    for s in sorted(stores_used):
        if not s:
            continue
        store_sql.append(
            f"INSERT OR IGNORE INTO stores (store_id,name,region,active) "
            f"VALUES ({q(s)},{q(STORE_NAMES.get(s, 'Liqo ' + s.title()))},'North India',1);"
        )
    summary["stores"] = len(store_sql)

    # ---------------------------------------------------------------- write SQL
    def write(name, header, stmts):
        with open(os.path.join(args.outdir, name), "w") as fh:
            fh.write("PRAGMA foreign_keys = ON;\n-- " + header + "\n")
            fh.write("\n".join(stmts) + "\n")

    write("01_stores.sql", "stores referenced by employees/inventory/sales", store_sql)
    write("02_employees.sql", "staff master (PII: names + mobiles)", emp_sql)
    write("04_customers.sql", "customers + brand prefs (PII)", cust_sql + [""] + bp_sql)
    write("05_sales.sql", "bills, lines, purchase events (PII)", sales_sql + [""] + item_sql + [""] + ev_sql)

    with open(os.path.join(args.outdir, "summary.json"), "w") as fh:
        json.dump(summary, fh, indent=2)
    print(json.dumps(summary, indent=2))


def rows_iter(it):
    """Yield rows, stopping at a fully-empty row (read_only sheets over-report dims)."""
    for r in it:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        yield r


if __name__ == "__main__":
    main()
